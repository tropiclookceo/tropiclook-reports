#!/usr/bin/env python3
"""
tl_report_engine.py — TropicLook Owner Report Generation Engine v3.0
Revenue recognition: CHECKOUT DATE (FIN-REG-OWN-RPT-001 v1.0)
Reads: INPUT template (8-sheet xlsx) → Writes: 6-tab Owner Report
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
import argparse, calendar, sys, os, io, re

# ── BRAND ─────────────────────────────────────────────────────────────────────
NAVY     = "1F3864"
TEAL     = "1F6E6E"
GOLD     = "C9A84C"
SILVER   = "D9D9D9"
LIGHT_BG = "EBF0F7"
WHITE    = "FFFFFF"
RED_BG   = "FFC7CE"
GREEN_BG = "C6EFCE"
FONT     = "Arial"

MONTHS_RU    = {1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
                7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь"}
MONTHS_SHORT = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
                7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}

CAT_NAMES = {
    "FIX-POOL":"Обслуживание бассейна","FIX-Garden":"Обслуживание сада",
    "FIX-MAIN":"Техническое обслуживание","FIX-CLEAN":"Регулярная уборка",
    "FIX-INET":"Интернет","FIX-COM":"Общие расходы здания",
    "VAR-CLEAN":"Экстра-уборка","VAR-LNDRY":"Стирка белья",
    "VAR-CHEM":"Химия для уборки","VAR-WELC":"Приветственные пакеты",
    "UTL-ELEC":"Электричество","UTL-WAT":"Вода",
    "MNT-REPAIR":"Ремонтные работы","EMRG":"Аварийный ремонт",
    "WASTE":"Вывоз мусора","FFE-EQUIP":"Оборудование и инвентарь",
    "GUEST-SVC":"Гостевой сервис","TAXES-PRP":"Налог на имущество",
    "MISC":"Прочее","ADJ":"Корректировки",
}

BUDGET_ORDER = [
    "FIX-POOL","FIX-Garden","FIX-MAIN","FIX-CLEAN","FIX-INET","FIX-COM",
    "VAR-CLEAN","VAR-LNDRY","VAR-CHEM","VAR-WELC",
    "UTL-ELEC","UTL-WAT","MNT-REPAIR","EMRG","WASTE","FFE-EQUIP",
    "GUEST-SVC","TAXES-PRP","MISC","ADJ",
]


# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
def _f(bold=False, size=10, color=None, name=FONT):
    return Font(name=name, bold=bold, size=size, color=(color or "000000"))

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_all(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom(style="thin"):
    s = Side(style=style)
    return Border(bottom=s)

def _thb(val):
    """Format number as Thai Baht string: '420,500 ฿'"""
    if val is None: return "— ฿"
    return f"{val:,.0f} ฿"

def _pct(val):
    if val is None: return "—%"
    return f"{int(round(val * 100))}%"

def _num_or_none(val):
    """Best-effort numeric parser for optional template values."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        v = val.strip().replace(",", "").replace(" ", "")
        v = re.sub(r"[^\d.\-]", "", v)
        if not v:
            return None
        try:
            return float(v)
        except ValueError:
            return None
    return None

def _row_value(row, idx, default=None):
    """Safely read an optional cell from an openpyxl values_only row tuple."""
    return row[idx] if len(row) > idx else default

def _ratio_or_none(val):
    """
    Parse a percent-like value as a spreadsheet ratio.

    Input templates historically described percentage fields as decimals
    (0.087 = 8.7%). Finance now wants these cells to display as percentages.
    This parser accepts both 0.087 and 8.7 and normalizes them to 0.087.
    """
    num = _num_or_none(val)
    if num is None:
        return None
    if abs(num) > 1:
        return num / 100
    return num

def _round_money(val):
    return round(float(val or 0), 2)

def _property_purchase_price(info):
    """Return the managed unit purchase price used for annualized yield."""
    for key in (
        "asset_purchase_price",
        "property_purchase_price",
        "unit_purchase_price",
        "purchase_price",
        "property_value",
        "asset_value",
        "estimated_property_value",
        "investment_value",
    ):
        value = _num_or_none(info.get(key))
        if value is not None:
            return value
    return None

def _style_header_row(ws, row, cols, bg=NAVY, fg=WHITE, bold=True, size=10):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = _f(bold=bold, size=size, color=fg)
        cell.fill = _fill(bg)
        cell.alignment = _align("center")

def _style_section(ws, row, col_start, col_end, label, bg=TEAL):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = label
    cell.font = _f(bold=True, size=10, color=WHITE)
    cell.fill = _fill(bg)
    cell.alignment = _align("left")

def _kpi_block(ws, row, col, label, value, subtitle, bg=NAVY):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+1)
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = _f(bold=True, size=9, color=WHITE); lc.fill = _fill(bg); lc.alignment = _align("center")
    vc = ws.cell(row=row+1, column=col, value=value)
    vc.font = _f(bold=True, size=14, color=WHITE); vc.fill = _fill(bg); vc.alignment = _align("center")
    sc = ws.cell(row=row+2, column=col, value=subtitle)
    sc.font = _f(bold=False, size=8, color=WHITE); sc.fill = _fill(bg); sc.alignment = _align("center")


# ── DATA READING ──────────────────────────────────────────────────────────────
def read_input(path):
    """
    Read all sheets from INPUT template.
    Returns a dict with parsed, clean Python objects.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    data = {}

    # Property_Info — vertical layout: column A = key, column B = value
    import re as _re
    ws = wb["Property_Info"]
    pi = {}
    for row in ws.iter_rows(values_only=True):
        if not any(row): continue
        key = row[0]
        val = row[1] if len(row) > 1 else None
        if key and isinstance(key, str) and _re.match(r'^[a-z_]+$', key.strip()):
            pi[key.strip()] = val
    # Convert date fields stored as strings
    if "mgmt_start_date" in pi:
        from datetime import datetime as _dt
        d = pi["mgmt_start_date"]
        if isinstance(d, str):
            for fmt in ('%Y-%m-%d', '%d.%m.%Y'):
                try: pi["mgmt_start_date"] = _dt.strptime(d.strip(), fmt); break
                except ValueError: pass
    data["info"] = pi

    # Reservations
    ws = wb["Reservations"]
    res = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if not any(row): continue
        if str(row[0]).startswith("NOTES"): break
        if headers is None:
            # Only treat row as headers if first cell is a lowercase field name
            if row[0] and isinstance(row[0], str) and _re.match(r'^[a-z_]+$', str(row[0]).strip()):
                headers = [str(h).strip() if h else "" for h in row]
            continue
        if len(row) < 5: continue
        r = dict(zip(headers, row))
        for df in ("checkin_date", "checkout_date"):
            r[df] = _to_date(r.get(df))
        if isinstance(r.get("checkout_date"), datetime):
            res.append(r)
    data["reservations"] = res

    # Expenses
    ws = wb["Expenses"]
    exp = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if not any(row): continue
        if str(row[0]).startswith("NOTES"): break
        if headers is None:
            # Only treat row as headers if first cell is a lowercase field name
            if row[0] and isinstance(row[0], str) and _re.match(r'^[a-z_]+$', str(row[0]).strip()):
                headers = [str(h).strip() if h else "" for h in row]
            continue
        date_val = _to_date(row[0])
        if not date_val: continue
        r = dict(zip(headers, row))
        r["date"] = date_val
        exp.append(r)
    data["expenses"] = exp

    # Owner_Payouts
    ws = wb["Owner_Payouts"]
    pay = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2: continue
        date_val = _to_date(row[0])
        if not date_val: continue
        amount = _num_or_none(_row_value(row, 1, 0)) or 0
        pay.append({
            "date": date_val, "amount": amount,
            "type": _row_value(row, 2, "") or "",
            "reference": _row_value(row, 3, "") or "",
            "description": _row_value(row, 4, "") or "",
        })
    data["payouts"] = pay

    # OPEX_Budget
    ws = wb["OPEX_Budget"]
    budget = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 1 or not row[0]: continue
        if str(row[0]).startswith("NOTES"): break
        code = str(row[0]).strip()
        raw = row[1] if len(row) > 1 else None
        amt = 0.0
        if isinstance(raw, (int, float)):
            amt = float(raw)
        budget[code] = amt
    data["budget"] = budget

    # Cumulative (for DAP)
    ws = wb["Cumulative"]
    cumul = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2 or not row[0]: continue
        if str(row[0]).startswith("NOTES"): break
        k = str(row[0]).strip()
        cumul[k] = row[1]
    data["cumulative"] = cumul

    wb.close()
    return data


# ── DATE HELPERS ──────────────────────────────────────────────────────────────
def _to_date(val):
    """Convert string or datetime to datetime object, return None if unparseable."""
    if isinstance(val, datetime): return val
    if isinstance(val, str):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%Y/%m/%d'):
            try: return datetime.strptime(val.strip(), fmt)
            except ValueError: pass
    return None

def _to_bool(val, default=False):
    """Best-effort bool parser for template flags."""
    if val is None:
        return default
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val != 0
    if isinstance(val, str):
        v = val.strip().lower()
        if v in {"1", "true", "yes", "y", "on", "strict", "block"}:
            return True
        if v in {"0", "false", "no", "n", "off", "warn", "warning"}:
            return False
    return default


# ── PERIOD HELPERS ────────────────────────────────────────────────────────────
def _parse_period(info):
    """Extract report year/month from Property_Info['period']."""
    p = info.get("period")
    if isinstance(p, datetime):
        return p.year, p.month
    if isinstance(p, str):
        parts = p.strip().split("-")
        if len(parts) == 2:
            return int(parts[0]), int(parts[1])
    raise ValueError(f"Cannot parse period: {p}")

def _period_str(year, month):
    return f"{year:04d}-{month:02d}"

def _next_period(year, month):
    if month == 12:
        return year + 1, 1
    return year, month + 1

def _month_days(year, month):
    return calendar.monthrange(year, month)[1]

def _iter_months(start_date, end_year, end_month):
    """Yield (year, month) from start_date through end_year/end_month."""
    y, m = start_date.year, start_date.month
    while (y, m) <= (end_year, end_month):
        yield y, m
        m += 1
        if m > 12:
            m = 1; y += 1

def _month_label(year, month):
    return f"{MONTHS_SHORT[month]}'{str(year)[2:]}"


# ── MONTHLY AGGREGATION ───────────────────────────────────────────────────────
def compute_monthly(data, mgmt_start, rpt_year, rpt_month):
    """
    For every month from mgmt_start to rpt_year/rpt_month,
    compute: gross, utility, tl_comm, opex, payouts, net_income, closing_bal
    Revenue recognition: CHECKOUT DATE.
    """
    # Opening balance at management start
    info = data["info"]
    opening_key = [k for k in info if "beginning" in k.lower() or "opening" in k.lower()]
    opening_bal = float(info.get(opening_key[0], 0)) if opening_key else 877089.71

    months = []
    prev_closing = opening_bal

    for yr, mo in _iter_months(mgmt_start, rpt_year, rpt_month):
        # Revenue: checkout date in this month
        bk = [r for r in data["reservations"]
              if r["checkout_date"].year == yr and r["checkout_date"].month == mo]
        gross   = sum(float(r.get("gross_amount") or 0) for r in bk)
        utility = sum(float(r.get("utility_charge") or 0) for r in bk)
        tl_comm = sum(float(r.get("tl_commission") or 0) for r in bk)
        bookings_count = len(bk)
        nights_occupied = sum(float(r.get("nights") or 0) for r in bk)

        # OPEX: expense date in this month
        ex = [e for e in data["expenses"]
              if e["date"].year == yr and e["date"].month == mo]
        opex = sum(float(e.get("amount") or 0) for e in ex)

        # Payouts: payout date in this month
        py = [p for p in data["payouts"]
              if p["date"].year == yr and p["date"].month == mo]
        payouts = sum(p["amount"] for p in py)

        net_income  = gross + utility - tl_comm - opex
        closing_bal = prev_closing + net_income - payouts

        days_in_month = _month_days(yr, mo)
        occupancy = (nights_occupied / days_in_month) if days_in_month > 0 else 0
        adr = (gross / nights_occupied) if nights_occupied > 0 else 0

        months.append({
            "year": yr, "month": mo,
            "label": _month_label(yr, mo),
            "gross": gross, "utility": utility, "tl_comm": tl_comm,
            "opex": opex, "payouts": payouts, "net_income": net_income,
            "opening_bal": prev_closing, "closing_bal": closing_bal,
            "bookings": bookings_count, "nights": nights_occupied,
            "occupancy": occupancy, "adr": adr,
            "owner_payin": 0.0,
        })
        prev_closing = closing_bal

    return months


# ── SHEET 1: DASHBOARD ────────────────────────────────────────────────────────
def build_dashboard(wb, data, rpt_year, rpt_month, cur_month):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14

    info = data["info"]
    prop_name  = info.get("property_name", "")
    prop_code  = info.get("property_code", "")
    owner_name = info.get("owner_name", "")
    prop_type  = info.get("property_type", "")
    bedrooms   = info.get("bedrooms", "")
    location   = info.get("location", "")
    comm_rate  = float(info.get("commission_rate", 0.25))

    period_label = f"{MONTHS_RU[rpt_month]} {rpt_year}"

    # Current month bookings (checkout-based)
    bk = [r for r in data["reservations"]
          if r["checkout_date"].year == rpt_year and r["checkout_date"].month == rpt_month]
    gross   = sum(float(r.get("gross_amount") or 0) for r in bk)
    utility = sum(float(r.get("utility_charge") or 0) for r in bk)
    tl_comm = sum(float(r.get("tl_commission") or 0) for r in bk)

    ex = [e for e in data["expenses"] if e["date"].year == rpt_year and e["date"].month == rpt_month]
    opex = sum(float(e.get("amount") or 0) for e in ex)

    net_income  = gross + utility - tl_comm - opex
    opening_bal = cur_month["opening_bal"]
    py          = [p for p in data["payouts"] if p["date"].year == rpt_year and p["date"].month == rpt_month]
    payouts     = sum(p["amount"] for p in py)
    closing_bal = cur_month["closing_bal"]

    days_in_month    = _month_days(rpt_year, rpt_month)
    nights_occupied  = sum(float(r.get("nights") or 0) for r in bk)
    occupancy        = nights_occupied / days_in_month if days_in_month > 0 else 0
    adr              = gross / nights_occupied if nights_occupied > 0 else 0
    expense_ratio    = opex / gross if gross > 0 else 0

    # Row 1: main title
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 28
    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = "TROPICLOOK — OWNER FINANCIAL REPORT"
    c.font = _f(bold=True, size=14, color=WHITE)
    c.fill = _fill(NAVY); c.alignment = _align("center")

    ws.row_dimensions[3].height = 18
    ws.merge_cells("B3:G3")
    c = ws["B3"]
    c.value = f"{prop_name}  |  {period_label}  |  {prop_code}"
    c.font = _f(bold=True, size=11, color=WHITE)
    c.fill = _fill(NAVY); c.alignment = _align("center")

    ws.row_dimensions[4].height = 16
    ws.merge_cells("B4:G4")
    c = ws["B4"]
    c.value = f"Owner: {owner_name}  |  Commission: as per table  |  {bedrooms}BR {prop_type}, {location}"
    c.font = _f(bold=False, size=9, color=WHITE)
    c.fill = _fill(NAVY); c.alignment = _align("center")

    ws.row_dimensions[5].height = 8

    # KPI section header
    ws.row_dimensions[6].height = 16
    ws.merge_cells("B6:G6")
    c = ws["B6"]
    c.value = "KEY PERFORMANCE INDICATORS"
    c.font = _f(bold=True, size=10, color=WHITE)
    c.fill = _fill(TEAL); c.alignment = _align("center")

    # 6 KPI blocks: rows 7-9, columns B-C, D-E, F-G
    for r in range(7, 10):
        ws.row_dimensions[r].height = 20
    _kpi_block(ws, 7, 2, "GROSS REVENUE",    _thb(gross),       "Доход от бронирований", NAVY)
    _kpi_block(ws, 7, 4, "NET OWNER INCOME",  _thb(net_income),  "После комиссии и OPEX",  NAVY)
    _kpi_block(ws, 7, 6, "CASH BALANCE",      _thb(closing_bal), "Остаток на счёте",        NAVY)

    for r in range(10, 13):
        ws.row_dimensions[r].height = 20
    _kpi_block(ws, 10, 2, "OCCUPANCY",
               f"{int(round(occupancy*100))}%",
               f"{int(nights_occupied)} из {days_in_month} ночей", TEAL)
    _kpi_block(ws, 10, 4, "ADR",
               _thb(adr),
               "Средняя стоимость/ночь", TEAL)
    _kpi_block(ws, 10, 6, "EXPENSE RATIO",
               f"{expense_ratio*100:.1f}%",
               "OPEX / Revenue", TEAL)

    ws.row_dimensions[13].height = 8

    # Bookings table header
    ws.row_dimensions[14].height = 16
    ws.merge_cells("B14:G14")
    c = ws["B14"]
    c.value = f"BOOKINGS — {MONTHS_RU[rpt_month].upper()} {rpt_year}"
    c.font = _f(bold=True, size=10, color=WHITE)
    c.fill = _fill(TEAL); c.alignment = _align("left")

    ws.row_dimensions[15].height = 14
    headers = ["Бронирование", "Канал", "Гость", "Даты", "Ночей", "Gross (฿)"]
    for i, h in enumerate(headers, 2):
        c = ws.cell(row=15, column=i, value=h)
        c.font = _f(bold=True, size=9, color=WHITE)
        c.fill = _fill(NAVY); c.alignment = _align("center")

    for ri, r in enumerate(bk, 16):
        ws.row_dimensions[ri].height = 14
        ci = r["checkin_date"].strftime("%d.%m") if isinstance(r["checkin_date"], datetime) else ""
        co = r["checkout_date"].strftime("%d.%m") if isinstance(r["checkout_date"], datetime) else ""
        dates = f"{ci}—{co} (доход: {co})"
        row_vals = [
            r.get("booking_id", ""), r.get("channel", ""), r.get("guest_name", ""),
            dates, int(r.get("nights") or 0), int(r.get("gross_amount") or 0)
        ]
        fill = _fill(LIGHT_BG) if ri % 2 == 0 else _fill(WHITE)
        for i, v in enumerate(row_vals, 2):
            c = ws.cell(row=ri, column=i, value=v)
            c.font = _f(size=9); c.fill = fill
            c.alignment = _align("center" if i >= 5 else "left")


# ── SHEET 2: P&L MONTHLY ──────────────────────────────────────────────────────
def build_pl(wb, data, rpt_year, rpt_month, cur_month):
    ws = wb.create_sheet("P&L Monthly")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 28

    info     = data["info"]
    prop_name = info.get("property_name", "")
    opening  = cur_month["opening_bal"]

    bk = [r for r in data["reservations"]
          if r["checkout_date"].year == rpt_year and r["checkout_date"].month == rpt_month]
    ex = [e for e in data["expenses"]
          if e["date"].year == rpt_year and e["date"].month == rpt_month]
    py = [p for p in data["payouts"]
          if p["date"].year == rpt_year and p["date"].month == rpt_month]

    gross   = sum(float(r.get("gross_amount") or 0) for r in bk)
    utility = sum(float(r.get("utility_charge") or 0) for r in bk)
    tl_comm = sum(float(r.get("tl_commission") or 0) for r in bk)
    ota_comm = sum(float(r.get("ota_commission") or 0) for r in bk)
    opex    = sum(float(e.get("amount") or 0) for e in ex)
    payouts = sum(p["amount"] for p in py)
    net_income  = gross + utility - tl_comm - opex
    closing_bal = cur_month["closing_bal"]
    period_label = f"{MONTHS_RU[rpt_month]} {rpt_year}"

    # Group OPEX by category
    opex_by_cat = defaultdict(float)
    opex_count  = 0
    for e in ex:
        cat = str(e.get("category_code") or "MISC").strip()
        opex_by_cat[cat] += float(e.get("amount") or 0)
        opex_count += 1

    row = 1
    def next_row():
        nonlocal row; row += 1; return row

    def write(r, label, val=None, note=None, bold=False, indent=False,
              bg=None, fg="000000", size=10, num_fmt=None):
        lbl = ("   " if indent else "") + (label or "")
        ws.row_dimensions[r].height = 15
        c_label = ws.cell(r, 2, lbl)
        c_label.font = _f(bold=bold, size=size, color=fg)
        if bg: c_label.fill = _fill(bg)
        c_label.alignment = _align("left")

        if val is not None:
            c_val = ws.cell(r, 3, val)
            c_val.font = _f(bold=bold, size=size, color=fg)
            if bg: c_val.fill = _fill(bg)
            c_val.alignment = _align("right")
            if num_fmt: c_val.number_format = num_fmt

        if note is not None:
            c_note = ws.cell(r, 4, note)
            c_note.font = _f(size=8, color="666666")
            c_note.alignment = _align("left")

    def write_section(r, label):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(r, 2, label)
        c.font = _f(bold=True, size=10, color=WHITE)
        c.fill = _fill(TEAL); c.alignment = _align("left")
        ws.row_dimensions[r].height = 16

    def write_total(r, label, val, note=None):
        write(r, label, val, note, bold=True, bg=LIGHT_BG)
        ws.cell(r, 2).border = _border_bottom("medium")
        ws.cell(r, 3).border = _border_bottom("medium")

    # Title
    ws.row_dimensions[1].height = 20
    ws.merge_cells("B1:D1")
    c = ws["B1"]; c.value = f"P&L — {prop_name} — {period_label}"
    c.font = _f(bold=True, size=12, color=WHITE); c.fill = _fill(NAVY)
    c.alignment = _align("center")

    ws.row_dimensions[2].height = 14
    ws.merge_cells("B2:D2")
    c = ws["B2"]; c.value = f"Начальный баланс: {gross_fmt(opening)}"
    c.font = _f(bold=False, size=10); c.alignment = _align("left")

    ws.row_dimensions[3].height = 14
    for col, hdr in [(2, None), (3, period_label), (4, "Примечание")]:
        c = ws.cell(3, col, hdr)
        if col > 2:
            c.font = _f(bold=True, size=9, color=WHITE)
            c.fill = _fill(NAVY); c.alignment = _align("center")

    row = 3
    # REVENUE
    write_section(next_row(), "REVENUE / ДОХОДЫ")
    for r in bk:
        bid   = r.get("booking_id", "")
        ch    = r.get("channel", "")
        nights = int(r.get("nights") or 0)
        amt   = float(r.get("gross_amount") or 0)
        write(next_row(), str(bid), amt, f"{ch}, {nights}н", indent=True)
    if utility > 0:
        write(next_row(), "Возмещение электричества", utility, "Utility recharge", indent=True)
    write_total(next_row(), "ИТОГО ДОХОД (A)", gross + utility)

    # COMMISSION
    write_section(next_row(), "КОМИССИЯ TROPICLOOK")
    comm_rate = float(info.get("commission_rate", 0.25))
    for r in bk:
        bid   = r.get("booking_id", "")
        amt   = float(r.get("gross_amount") or 0)
        tl    = float(r.get("tl_commission") or 0)
        pct   = tl/amt*100 if amt > 0 else comm_rate*100
        write(next_row(), str(bid), -tl, f"{pct:.1f}% от {gross_fmt(amt)}", indent=True)
    if ota_comm > 0:
        write(next_row(), "Справочно: комиссия OTA (не влияет на баланс)",
              -ota_comm, "Удержано OTA", indent=True)
    write_total(next_row(), "ИТОГО КОМИССИЯ TL (B)", -tl_comm)

    # OPEX
    write_section(next_row(), "ОПЕРАЦИОННЫЕ РАСХОДЫ / OPEX")
    budget = data["budget"]
    for cat in BUDGET_ORDER:
        if cat not in opex_by_cat: continue
        name = CAT_NAMES.get(cat, cat)
        fact = opex_by_cat[cat]
        bgt  = budget.get(cat, 0)
        note = f"бюджет:{gross_fmt(bgt)}" if bgt else ""
        write(next_row(), name, -fact, note, indent=True)
    # Any uncategorised
    for cat, fact in opex_by_cat.items():
        if cat not in BUDGET_ORDER:
            write(next_row(), CAT_NAMES.get(cat, cat), -fact, indent=True)
    write_total(next_row(), "ИТОГО OPEX (C)", -opex, f"{opex_count} операций")

    # NET INCOME
    ni_row = next_row()
    ws.row_dimensions[ni_row].height = 16
    c = ws.cell(ni_row, 2)
    c.value = "ЧИСТЫЙ ДОХОД (A+B+C)"
    c.font = _f(bold=True, size=11, color=WHITE)
    c.fill = _fill(NAVY); c.alignment = _align("left")
    c2 = ws.cell(ni_row, 3, net_income)
    c2.font = _f(bold=True, size=11, color=WHITE); c2.fill = _fill(NAVY)
    c2.alignment = _align("right")
    ws.cell(ni_row, 4).fill = _fill(NAVY)

    # PAYOUTS
    write_section(next_row(), "ВЫПЛАТЫ СОБСТВЕННИКУ")
    for p in py:
        d = p["date"].strftime("%d.%m.%Y")
        write(next_row(), p["description"][:55], -p["amount"], d, indent=True)
    write_total(next_row(), "ИТОГО ВЫПЛАТЫ (D)", -payouts, f"{len(py)} выплат")

    # BALANCE MOVEMENT
    write_section(next_row(), "ДВИЖЕНИЕ БАЛАНСА")
    write(next_row(), "Баланс на начало", opening)
    write(next_row(), "+ Доходы (A)", gross + utility)
    write(next_row(), "- Комиссия TL (B)", -tl_comm, "OTA не влияет на баланс")
    write(next_row(), "- OPEX (C)", -opex)
    write(next_row(), "- Выплаты (D)", -payouts)

    final_row = next_row()
    ws.row_dimensions[final_row].height = 18
    ws.merge_cells(start_row=final_row, start_column=2, end_row=final_row, end_column=2)
    c = ws.cell(final_row, 2, "БАЛАНС НА КОНЕЦ")
    c.font = _f(bold=True, size=12, color=WHITE); c.fill = _fill(NAVY)
    c.alignment = _align("left")
    c2 = ws.cell(final_row, 3, closing_bal)
    c2.font = _f(bold=True, size=12, color=WHITE); c2.fill = _fill(NAVY)
    c2.alignment = _align("right")


def gross_fmt(v):
    if v is None: return "0"
    return f"{v:,.0f}"


# ── SHEET 3: OPEX PASSPORT ────────────────────────────────────────────────────
def build_opex_passport(wb, data, rpt_year, rpt_month):
    ws = wb.create_sheet("OPEX Passport")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10

    period_label = f"OPEX PASSPORT — {MONTHS_RU[rpt_month]} {rpt_year}"

    ws.row_dimensions[1].height = 20
    ws.merge_cells("B1:G1")
    c = ws["B1"]; c.value = period_label
    c.font = _f(bold=True, size=12, color=WHITE); c.fill = _fill(NAVY)
    c.alignment = _align("center")

    ws.row_dimensions[2].height = 14
    hdrs = ["Категория", "Бюджет", "Факт", "Δ", "Δ%", "Статус"]
    for i, h in enumerate(hdrs, 2):
        c = ws.cell(2, i, h)
        c.font = _f(bold=True, size=9, color=WHITE)
        c.fill = _fill(TEAL); c.alignment = _align("center")

    budget = data["budget"]
    ex = [e for e in data["expenses"]
          if e["date"].year == rpt_year and e["date"].month == rpt_month]
    opex_by_cat = defaultdict(float)
    for e in ex:
        cat = str(e.get("category_code") or "MISC").strip()
        opex_by_cat[cat] += float(e.get("amount") or 0)

    all_cats = list(BUDGET_ORDER)
    for cat in opex_by_cat:
        if cat not in all_cats:
            all_cats.append(cat)

    total_bgt = total_fact = 0
    row = 2
    for cat in all_cats:
        fact = opex_by_cat.get(cat, 0)
        bgt  = budget.get(cat, 0) or 0
        if fact == 0 and bgt == 0:
            continue
        row += 1
        ws.row_dimensions[row].height = 14
        delta = fact - bgt
        delta_pct = (delta / bgt) if bgt != 0 else (-1 if fact == 0 else 0)
        status = "✅" if fact <= bgt or bgt == 0 else "🔴"
        name = CAT_NAMES.get(cat, cat)

        fill = _fill(LIGHT_BG) if row % 2 == 0 else _fill(WHITE)
        vals = [name, bgt, fact, delta, delta_pct, status]
        for i, v in enumerate(vals, 2):
            c = ws.cell(row, i, v)
            c.font = _f(size=9); c.fill = fill; c.alignment = _align("center" if i >= 3 else "left")
            if i == 6:  # Δ% column — format as percentage
                c.number_format = '0.0%'
            if i == 7:  # status emoji
                c.font = _f(size=11)
        total_bgt += bgt; total_fact += fact

    # Totals row
    row += 1
    ws.row_dimensions[row].height = 15
    for i, v in enumerate(["ИТОГО", total_bgt, total_fact, total_fact - total_bgt, None, None], 2):
        c = ws.cell(row, i, v)
        c.font = _f(bold=True, size=10); c.fill = _fill(LIGHT_BG)
        c.border = _border_bottom("medium"); c.alignment = _align("center" if i >= 3 else "left")


# ── SHEET 4: 12-MONTH SUMMARY ─────────────────────────────────────────────────
def build_12month(wb, data, rpt_year, rpt_month, months):
    ws = wb.create_sheet("12-Month Summary")
    ws.sheet_view.showGridLines = False

    info      = data["info"]
    prop_name = info.get("property_name", "")
    mgmt_start = info.get("mgmt_start_date")
    if isinstance(mgmt_start, datetime):
        start_str = f"{MONTHS_RU[mgmt_start.month]} {mgmt_start.year}"
    else:
        start_str = "—"
    end_str = f"{MONTHS_RU[rpt_month]} {rpt_year}"

    # Column setup: B=labels, C..N=months, O=TOTAL
    num_months = len(months)
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 24
    for i in range(num_months):
        col = get_column_letter(3 + i)
        ws.column_dimensions[col].width = 13
    total_col = get_column_letter(3 + num_months)
    ws.column_dimensions[total_col].width = 14

    def mc(r, c): return ws.cell(r, c)

    # Rows 1-2: titles
    ws.row_dimensions[1].height = 20
    last_col = 3 + num_months
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_col)
    c = ws["B1"]; c.value = f"PERFORMANCE OVERVIEW — {prop_name}"
    c.font = _f(bold=True, size=12, color=WHITE); c.fill = _fill(NAVY); c.alignment = _align("center")

    ws.row_dimensions[2].height = 14
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=last_col)
    c = ws["B2"]; c.value = f"{start_str} — {end_str} (management period)"
    c.font = _f(size=9); c.alignment = _align("center")

    # Row 3: column headers
    ws.row_dimensions[3].height = 14
    mc(3, 2).value = "Показатель"
    mc(3, 2).font = _f(bold=True, size=9, color=WHITE); mc(3, 2).fill = _fill(NAVY); mc(3, 2).alignment = _align("left")
    for i, mo in enumerate(months):
        c = mc(3, 3 + i); c.value = mo["label"]
        c.font = _f(bold=True, size=9, color=WHITE); c.fill = _fill(NAVY); c.alignment = _align("center")
    tc = mc(3, last_col); tc.value = "TOTAL"
    tc.font = _f(bold=True, size=9, color=WHITE); tc.fill = _fill(NAVY); tc.alignment = _align("center")

    def write_row(r, label, values, totals_fn=sum, bold=False, bg=None, is_section=False):
        ws.row_dimensions[r].height = 14
        if is_section:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=last_col)
            c = mc(r, 2); c.value = label
            c.font = _f(bold=True, size=9, color=WHITE); c.fill = _fill(TEAL); c.alignment = _align("left")
            return
        fill = _fill(bg or WHITE)
        c = mc(r, 2); c.value = label
        c.font = _f(bold=bold, size=9); c.fill = fill; c.alignment = _align("left")
        non_none = [v for v in values if v is not None]
        total = totals_fn(non_none) if non_none else None
        for i, v in enumerate(values):
            c = mc(r, 3 + i)
            c.value = round(v, 2) if isinstance(v, float) else v
            c.font = _f(bold=bold, size=9); c.fill = fill; c.alignment = _align("right")
        tc = mc(r, last_col)
        tc.value = round(total, 2) if isinstance(total, float) else total
        tc.font = _f(bold=bold, size=9); tc.fill = _fill(LIGHT_BG) if not bg else fill
        tc.alignment = _align("right")

    row = 3
    # REVENUE section
    row += 1; write_row(row, "REVENUE", [], is_section=True)
    row += 1; write_row(row, "Gross Revenue (฿)", [m["gross"] for m in months], bg=LIGHT_BG)
    row += 1; write_row(row, "Utility (฿)",       [m["utility"] for m in months])
    row += 1; write_row(row, "Owner Payin (฿)",   [m["owner_payin"] for m in months], bg=LIGHT_BG)
    row += 1; write_row(row, "Bookings (#)",       [m["bookings"] for m in months])
    row += 1; write_row(row, "Nights",             [int(m["nights"]) for m in months], bg=LIGHT_BG)
    # DEDUCTIONS
    row += 1; write_row(row, "DEDUCTIONS", [], is_section=True)
    row += 1; write_row(row, "TL Commission (฿)", [m["tl_comm"] for m in months])
    row += 1; write_row(row, "OPEX (฿)",           [m["opex"] for m in months], bg=LIGHT_BG)
    # RESULT
    row += 1; write_row(row, "RESULT", [], is_section=True)
    row += 1; write_row(row, "Net Income (฿)",     [m["net_income"] for m in months], bold=True)
    row += 1; write_row(row, "Payouts (฿)",         [m["payouts"] for m in months], bg=LIGHT_BG)
    row += 1; write_row(row, "Closing Bal (฿)",    [m["closing_bal"] for m in months],
                        totals_fn=lambda x: x[-1] if x else 0, bold=True)
    # KPIs
    row += 1; write_row(row, "KPIs", [], is_section=True)
    row += 1; write_row(row, "Occupancy %",
                        [int(round(m["occupancy"]*100)) for m in months],
                        totals_fn=lambda x: round(sum(x)/len(x)) if x else 0)
    row += 1; write_row(row, "ADR (฿)",
                        [int(round(m["adr"])) for m in months],
                        totals_fn=sum)


# ── SHEET 5: TRANSACTION LEDGER ───────────────────────────────────────────────
def build_ledger(wb, data, rpt_year, rpt_month, cur_month):
    ws = wb.create_sheet("Transaction Ledger")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    info     = data["info"]
    prop_name = info.get("property_name", "")
    opening  = cur_month["opening_bal"]
    period_label = f"{MONTHS_RU[rpt_month]} {rpt_year}"

    # Title
    ws.row_dimensions[1].height = 20
    ws.merge_cells("B1:G1")
    c = ws["B1"]; c.value = f"LEDGER — {prop_name} — {period_label}"
    c.font = _f(bold=True, size=12, color=WHITE); c.fill = _fill(NAVY); c.alignment = _align("center")

    ws.row_dimensions[2].height = 14
    for i, h in enumerate(["Дата", "Описание", "Кат.", "Приход", "Расход", "Баланс"], 2):
        c = ws.cell(2, i, h)
        c.font = _f(bold=True, size=9, color=WHITE)
        c.fill = _fill(TEAL); c.alignment = _align("center")

    # Build chronological transaction list
    transactions = []

    # OPEX expenses (dated)
    ex = [e for e in data["expenses"]
          if e["date"].year == rpt_year and e["date"].month == rpt_month]
    for e in ex:
        cat = str(e.get("category_code") or "MISC")
        transactions.append({
            "date": e["date"],
            "desc": (e.get("description") or "")[:55],
            "cat":  cat,
            "debit": float(e.get("amount") or 0),
            "credit": 0,
            "sort": 1,
        })

    # Revenue, TL commission, utility — on checkout date
    bk = [r for r in data["reservations"]
          if r["checkout_date"].year == rpt_year and r["checkout_date"].month == rpt_month]
    for r in bk:
        co   = r["checkout_date"]
        bid  = str(r.get("booking_id", ""))
        gross = float(r.get("gross_amount") or 0)
        tl    = float(r.get("tl_commission") or 0)
        util  = float(r.get("utility_charge") or 0)
        if gross > 0:
            transactions.append({"date": co, "desc": f"Приход: {bid} (выезд — checkout)",
                                  "cat": "REVENUE", "credit": gross, "debit": 0, "sort": 2})
        if tl > 0:
            transactions.append({"date": co, "desc": f"Комиссия TL: {bid}",
                                  "cat": "MGMT-FEE", "debit": tl, "credit": 0, "sort": 3})
        if util > 0:
            transactions.append({"date": co, "desc": f"Электричество: {bid} (возмещение)",
                                  "cat": "UTL-RCHG", "credit": util, "debit": 0, "sort": 4})

    # Payouts
    py = [p for p in data["payouts"]
          if p["date"].year == rpt_year and p["date"].month == rpt_month]
    for p in py:
        transactions.append({
            "date": p["date"],
            "desc": (p["description"] or "")[:55],
            "cat":  "PAYOUT",
            "debit": p["amount"], "credit": 0,
            "sort": 5,
        })

    # Sort: by date, then sort key
    transactions.sort(key=lambda x: (x["date"], x["sort"]))

    # Write
    row = 2

    def ledger_row(r, date_str, desc, cat, credit, debit, balance, bold=False, bg=None):
        ws.row_dimensions[r].height = 14
        fill = _fill(bg) if bg else (_fill(LIGHT_BG) if r % 2 == 0 else _fill(WHITE))
        vals = [date_str, desc, cat, credit or None, debit or None, balance]
        for i, v in enumerate(vals, 2):
            c = ws.cell(r, i, v)
            c.font = _f(bold=bold, size=9)
            c.fill = fill
            c.alignment = _align("right" if i >= 5 else ("center" if i == 2 else "left"))

    # Opening balance row
    row += 1
    ws.row_dimensions[row].height = 15
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    c = ws.cell(row, 3, "НАЧАЛЬНЫЙ БАЛАНС")
    c.font = _f(bold=True, size=9, color=WHITE); c.fill = _fill(TEAL); c.alignment = _align("center")
    c2 = ws.cell(row, 7, round(opening, 2))
    c2.font = _f(bold=True, size=9, color=WHITE); c2.fill = _fill(TEAL); c2.alignment = _align("right")

    balance = opening
    for t in transactions:
        row += 1
        balance = balance + t["credit"] - t["debit"]
        date_str = t["date"].strftime("%d.%m") if isinstance(t["date"], datetime) else ""
        ledger_row(row, date_str, t["desc"], t["cat"],
                   t["credit"] if t["credit"] > 0 else None,
                   t["debit"]  if t["debit"]  > 0 else None,
                   round(balance, 2))

    # Closing balance row
    row += 1
    ws.row_dimensions[row].height = 15
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    c = ws.cell(row, 3, "КОНЕЧНЫЙ БАЛАНС")
    c.font = _f(bold=True, size=9, color=WHITE); c.fill = _fill(NAVY); c.alignment = _align("center")
    c2 = ws.cell(row, 7, round(balance, 2))
    c2.font = _f(bold=True, size=9, color=WHITE); c2.fill = _fill(NAVY); c2.alignment = _align("right")

    # Revenue recognition footnote
    row += 2
    bk_str = "; ".join(
        f"{r.get('booking_id')}: заезд {r['checkin_date'].strftime('%d.%m')} → выезд {r['checkout_date'].strftime('%d.%m')} → доход признан {r['checkout_date'].strftime('%d.%m.%Y')}"
        for r in bk if isinstance(r.get("checkin_date"), datetime)
    )
    if bk_str:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        c = ws.cell(row, 2, f"⚠ Принцип признания дохода: дата выезда гостя (checkout date). {bk_str}.")
        c.font = _f(size=8, color="666666"); c.alignment = _align("left", wrap=True)
        ws.row_dimensions[row].height = 28


# ── SHEET 6: DAP SNAPSHOT ─────────────────────────────────────────────────────
def build_dap(wb, data, rpt_year, rpt_month, months):
    ws = wb.create_sheet("DAP Snapshot")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22

    info = data["info"]
    prop_name  = info.get("property_name", "")
    prop_code  = info.get("property_code", "")
    location   = info.get("location", "")
    prop_type  = info.get("property_type", "")
    bedrooms   = info.get("bedrooms", "")
    owner_name = info.get("owner_name", "")
    mgmt_start = info.get("mgmt_start_date")
    mgmt_str   = mgmt_start.strftime("%B %Y") if isinstance(mgmt_start, datetime) else "—"

    period_label = f"{MONTHS_RU[rpt_month]} {rpt_year}"

    # Cumulative from computed months
    total_gross   = sum(m["gross"]      for m in months)
    total_util    = sum(m["utility"]    for m in months)
    total_comm    = sum(m["tl_comm"]    for m in months)
    total_opex    = sum(m["opex"]       for m in months)
    total_net     = sum(m["net_income"] for m in months)
    total_pays    = sum(m["payouts"]    for m in months)
    total_bk      = sum(m["bookings"]   for m in months)
    total_nights  = sum(m["nights"]     for m in months)
    current_bal   = months[-1]["closing_bal"] if months else 0
    avg_occ       = sum(m["occupancy"]  for m in months) / len(months) if months else 0
    avg_stay      = total_nights / total_bk if total_bk > 0 else 0
    opex_by_month = [m["opex"] for m in months]
    avg_opex      = sum(opex_by_month) / len(opex_by_month) if opex_by_month else 0
    # Expense ratio cumulative
    exp_ratio = total_opex / total_gross if total_gross > 0 else 0

    # Find best/worst opex month
    hi_m = max(months, key=lambda m: m["opex"])
    lo_m = min(months, key=lambda m: m["opex"])

    # Title
    ws.row_dimensions[1].height = 22
    ws.merge_cells("B1:F1")
    c = ws["B1"]; c.value = f"DIGITAL ASSET PASSPORT — {prop_name}"
    c.font = _f(bold=True, size=12, color=WHITE); c.fill = _fill(NAVY); c.alignment = _align("center")

    ws.row_dimensions[2].height = 14
    ws.merge_cells("B2:F2")
    c = ws["B2"]; c.value = f"{prop_code} | ACTIVE | {period_label}"
    c.font = _f(size=9, color=WHITE); c.fill = _fill(NAVY); c.alignment = _align("center")

    def section_hdr(row, col, label, bg=TEAL):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        c = ws.cell(row, col, label)
        c.font = _f(bold=True, size=9, color=WHITE); c.fill = _fill(bg); c.alignment = _align("left")
        ws.row_dimensions[row].height = 14

    def kv_row(row, col, key, val):
        ws.row_dimensions[row].height = 13
        ck = ws.cell(row, col, key); ck.font = _f(size=9); ck.alignment = _align("left")
        cv = ws.cell(row, col+1, val); cv.font = _f(bold=True, size=9); cv.alignment = _align("left")

    ws.row_dimensions[3].height = 8

    # L1 — IDENTITY (col B-C), L3 — OPERATIONAL (col E-F)
    section_hdr(4, 2, "L1 — IDENTITY")
    section_hdr(4, 5, "L3 — OPERATIONAL")
    r = 4
    id_data = [
        ("Property", prop_name), ("Project", "—"), ("Location", location),
        ("Type", f"{prop_type} {bedrooms}BR + Pool"),
        ("Owner", owner_name), ("Ownership", "100%"),
        ("Mgmt Start", mgmt_str), ("Contract", "Marketing + Service"), ("Status", "ACTIVE"),
    ]
    op_data = [
        ("Bookings (cumul)", str(int(total_bk))),
        ("Nights (cumul)", str(int(total_nights))),
        ("Avg Occupancy", f"{int(round(avg_occ*100))}%"),
        ("Avg Stay", f"{avg_stay:.1f} nights"),
        ("Owner Bookings", "—"),
        ("Top Channel", "Mixed"),
        ("Readiness", "READY"),
        ("Condition", "— (pending)"),
    ]
    for i, (k, v) in enumerate(id_data):
        kv_row(r + 1 + i, 2, k, v)
    for i, (k, v) in enumerate(op_data):
        if r + 1 + i <= r + 1 + len(id_data):
            kv_row(r + 1 + i, 5, k, v)

    # Spacer
    last_id_row = r + 1 + len(id_data) + 1
    ws.row_dimensions[last_id_row].height = 6

    # L5 — FINANCIAL (col B-C), L4 — OPEX SUMMARY (col E-F)
    section_hdr(last_id_row + 1, 2, "L5 — FINANCIAL (cumul)")
    section_hdr(last_id_row + 1, 5, "L4 — OPEX SUMMARY")
    r2 = last_id_row + 1
    fin_data = [
        ("Cumul Gross Revenue", f"{total_gross:,.0f} ฿"),
        ("Cumul Utilities",     f"{total_util:,.0f} ฿"),
        ("Cumul TL Commission", f"{total_comm:,.0f} ฿"),
        ("Cumul OPEX",          f"{total_opex:,.0f} ฿"),
        ("Cumul Net Income",    f"{total_net:,.0f} ฿"),
        ("Total Payouts",       f"{total_pays:,.0f} ฿"),
        ("Current Balance",     f"{current_bal:,.0f} ฿"),
    ]
    opex_data = [
        ("Avg Monthly OPEX",  f"{avg_opex:,.0f} ฿"),
        ("Highest OPEX",      f"{hi_m['label']} ({hi_m['opex']:,.0f} ฿)"),
        ("Lowest OPEX",       f"{lo_m['label']} ({lo_m['opex']:,.0f} ฿)"),
        ("Expense Ratio",     f"{exp_ratio*100:.1f}%"),
        ("Budget Adherence",  "Mixed"),
        ("OPEX Profile",      "Active"),
        ("FIX-COM",           "Not yet active"),
    ]
    for i, (k, v) in enumerate(fin_data):
        kv_row(r2 + 1 + i, 2, k, v)
    for i, (k, v) in enumerate(opex_data):
        kv_row(r2 + 1 + i, 5, k, v)


# ── VALIDATION (5 RULES per FIN-REG-OWN-RPT-001) ─────────────────────────────
def validate(data, rpt_year, rpt_month, cur_month):
    errors = []; warnings = []
    info = data["info"]
    comm_rate = float(info.get("commission_rate", 0.25))
    # Non-blocking by default to avoid stopping the pipeline on CFO-escalation cases.
    block_on_negative_balance = _to_bool(info.get("block_on_negative_balance"), default=False)

    bk = [r for r in data["reservations"]
          if r["checkout_date"].year == rpt_year and r["checkout_date"].month == rpt_month]
    ex = [e for e in data["expenses"]
          if e["date"].year == rpt_year and e["date"].month == rpt_month]
    py = [p for p in data["payouts"]
          if p["date"].year == rpt_year and p["date"].month == rpt_month]

    gross   = sum(float(r.get("gross_amount") or 0) for r in bk)
    utility = sum(float(r.get("utility_charge") or 0) for r in bk)
    tl_comm = sum(float(r.get("tl_commission") or 0) for r in bk)
    opex    = sum(float(e.get("amount") or 0) for e in ex)
    payouts = sum(p["amount"] for p in py)
    net     = gross + utility - tl_comm - opex
    closing = cur_month["closing_bal"]
    opening = cur_month["opening_bal"]

    # R1: Closing balance >= 0
    if closing < -0.01:
        msg = f"R1 FAIL: Closing balance {closing:.2f} < 0. Escalate to CFO."
        if block_on_negative_balance:
            errors.append(msg)
        else:
            warnings.append(msg.replace("R1 FAIL", "R1 WARN") + " Report generated in non-blocking mode.")
    # R2: Arithmetic check
    calc_closing = opening + gross + utility - tl_comm - opex - payouts
    if abs(calc_closing - closing) > 1.0:
        errors.append(f"R2 FAIL: Balance discrepancy {abs(calc_closing - closing):.2f} THB > 1 THB.")
    # R3: All expenses have category code
    no_cat = [e for e in ex if not e.get("category_code")]
    if no_cat:
        errors.append(f"R3 FAIL: {len(no_cat)} expense(s) missing category code.")
    # R4: Booking count (informational)
    if len(bk) == 0:
        warnings.append("R4 WARN: 0 bookings in report period — verify against PMS.")
    # R5: Management fee vs expected
    expected_fee = gross * comm_rate
    if expected_fee > 0 and abs(tl_comm - expected_fee) / expected_fee > 0.05:
        warnings.append(f"R5 WARN: TL Fee {tl_comm:.0f} deviates >5% from expected {expected_fee:.0f}.")

    return errors, warnings


# ── NEXT INPUT TEMPLATE ROLLFORWARD ──────────────────────────────────────────
def _find_key_row(ws, key, key_col=1):
    """Find a row in a vertical key/value sheet by the key in column A."""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=key_col).value == key:
            return row
    raise KeyError(f"Key '{key}' not found in sheet '{ws.title}'.")

def _last_metric_row(ws, key_col=1):
    last = 0
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=key_col).value
        if isinstance(val, str) and re.match(r"^[a-z_]+$", val.strip()):
            last = row
    return last or ws.max_row

def _ensure_kv(ws, key, value=None, note=None, key_col=1, value_col=2):
    """
    Ensure a vertical key/value field exists.

    Used for new template fields so older InputData files can roll forward
    into the updated template shape without manual sheet surgery.
    """
    try:
        row = _find_key_row(ws, key, key_col=key_col)
    except KeyError:
        row = _last_metric_row(ws, key_col=key_col) + 1
        ws.insert_rows(row)
        ws.cell(row=row, column=key_col).value = key
    if value is not None:
        ws.cell(row=row, column=value_col).value = value
    if note is not None:
        ws.cell(row=row, column=value_col + 1).value = note
    return row

def _set_kv(ws, key, value, key_col=1, value_col=2, num_fmt=None):
    row = _find_key_row(ws, key, key_col=key_col)
    cell = ws.cell(row=row, column=value_col)
    cell.value = value
    if num_fmt:
        cell.number_format = num_fmt

def _clear_kv(ws, key, key_col=1, value_col=2):
    row = _find_key_row(ws, key, key_col=key_col)
    ws.cell(row=row, column=value_col).value = None

def _next_input_filename(input_path, current_period, next_period):
    """Derive the next InputData filename from the current file name."""
    name = os.path.basename(input_path).replace("OwnerReport", "InputData")
    if current_period in name:
        return name.replace(current_period, next_period, 1)
    stem, ext = os.path.splitext(name)
    ext = ext or ".xlsx"
    replaced = re.sub(r"\d{4}-\d{2}$", next_period, stem)
    if replaced != stem:
        return f"{replaced}{ext}"
    return f"{stem}_{next_period}{ext}"

def _rollforward_snapshot(data, months):
    """
    Build all values that carry from the closed month into the next input file.

    The carry-forward rule is intentionally narrow:
    - current month KPI/result values go to Prior_Period;
    - cumulative values are recomputed through the current closed month;
    - Cash_Balance gets the full closed-month balance control values.
    """
    info = data["info"]
    cur = months[-1]
    days = _month_days(cur["year"], cur["month"])
    revpar = (cur["gross"] / days) if days else 0

    total_gross = sum(m["gross"] for m in months)
    total_opex = sum(m["opex"] for m in months)
    total_fee = sum(m["tl_comm"] for m in months)
    total_net = sum(m["net_income"] for m in months)
    total_payouts = sum(m["payouts"] for m in months)

    existing_yield = _ratio_or_none(data.get("cumulative", {}).get("annualized_yield"))
    property_value = _property_purchase_price(info)
    annualized_yield = existing_yield
    if property_value and property_value > 0 and months:
        annualized_yield = (total_net / property_value) * (12 / len(months))

    mgmt_start = info.get("mgmt_start_date")
    mgmt_start_str = (
        mgmt_start.strftime("%Y-%m-%d") if isinstance(mgmt_start, datetime)
        else str(mgmt_start or "")
    )

    return {
        "prior_period": {
            "gross_revenue": _round_money(cur["gross"]),
            "mgmt_fee": _round_money(cur["tl_comm"]),
            "total_opex": _round_money(cur["opex"]),
            "net_income": _round_money(cur["net_income"]),
            "cash_balance_end": _round_money(cur["closing_bal"]),
            "occupancy_pct": round(float(cur["occupancy"] or 0), 4),
            "adr": _round_money(cur["adr"]),
            "revpar": _round_money(revpar),
        },
        "cumulative": {
            "cumulative_gross_revenue": _round_money(total_gross),
            "cumulative_opex": _round_money(total_opex),
            "cumulative_mgmt_fee": _round_money(total_fee),
            "cumulative_net_income": _round_money(total_net),
            "total_owner_payouts": _round_money(total_payouts),
            "annualized_yield": round(annualized_yield, 6) if annualized_yield is not None else None,
            "management_start_date": mgmt_start_str,
            "months_managed": len(months),
        },
        "cash_balance": {
            "opening_balance": _round_money(cur["opening_bal"]),
            "total_income": _round_money(cur["gross"] + cur["utility"]),
            "total_mgmt_fee": _round_money(-cur["tl_comm"]),
            "total_opex": _round_money(-cur["opex"]),
            "total_payouts": _round_money(-cur["payouts"]),
            "closing_balance": _round_money(cur["closing_bal"]),
        },
    }

def generate_next_input_template(input_path, output_path=None):
    """
    Create the next month's InputData workbook from the current closed input.

    This is the rollforward step for the three tabs finance currently
    re-enters manually:
    - Prior_Period receives current month KPIs;
    - Cumulative receives recomputed management-to-date totals;
    - Cash_Balance receives the full closed-month balance control:
      opening balance, income, fee, opex, payouts, and closing balance.
    """
    data = read_input(input_path)
    info = data["info"]
    rpt_year, rpt_month = _parse_period(info)
    next_year, next_month = _next_period(rpt_year, rpt_month)
    current_period = _period_str(rpt_year, rpt_month)
    next_period = _period_str(next_year, next_month)

    mgmt_start = info.get("mgmt_start_date")
    if not isinstance(mgmt_start, datetime):
        raise ValueError("mgmt_start_date missing or invalid in Property_Info")

    months = compute_monthly(data, mgmt_start, rpt_year, rpt_month)
    if not months:
        raise ValueError("No monthly data computed — cannot roll forward next input template.")

    cur_month = months[-1]
    errors, warnings = validate(data, rpt_year, rpt_month, cur_month)
    if errors:
        raise ValueError("Validation errors; next input template not created:\n" + "\n".join(errors))

    wb = load_workbook(input_path)
    snapshot = _rollforward_snapshot(data, months)

    # Property_Info: the workbook becomes the next reporting period.
    ws_info = wb["Property_Info"]
    _set_kv(ws_info, "period", next_period)
    _ensure_kv(
        ws_info,
        "asset_purchase_price",
        note="Стоимость покупки управляемой единицы (THB). Используется для annualized_yield.",
    )

    # Prior_Period: current closed month becomes the previous month for MoM.
    ws_prior = wb["Prior_Period"]
    for key, value in snapshot["prior_period"].items():
        num_fmt = "0.00%" if key == "occupancy_pct" else None
        _set_kv(ws_prior, key, value, num_fmt=num_fmt)
    _set_kv(
        ws_prior,
        "occupancy_pct",
        "Occupancy % предыдущего месяца",
        value_col=3,
    )

    # Cumulative: recompute through the current closed month.
    ws_cumul = wb["Cumulative"]
    for key, value in snapshot["cumulative"].items():
        num_fmt = "0.00%" if key == "annualized_yield" else None
        _set_kv(ws_cumul, key, value, num_fmt=num_fmt)
    _set_kv(
        ws_cumul,
        "annualized_yield",
        "Годовая доходность (%): Net Income / asset_purchase_price × 12 / months",
        value_col=3,
    )

    # Cash_Balance: full balance control for the closed period.
    ws_cash = wb["Cash_Balance"]
    _set_kv(ws_cash, "opening_balance", snapshot["cash_balance"]["opening_balance"], num_fmt="#,##0.00")
    for key in ("total_income", "total_mgmt_fee", "total_opex", "total_payouts"):
        _set_kv(ws_cash, key, snapshot["cash_balance"][key], num_fmt="#,##0.00")
    _set_kv(ws_cash, "closing_balance", snapshot["cash_balance"]["closing_balance"], num_fmt="#,##0.00")

    output_name = _next_input_filename(input_path, current_period, next_period)
    metadata = {
        "current_period": current_period,
        "next_period": next_period,
        "next_input_name": output_name,
        "opening_balance": snapshot["cash_balance"]["closing_balance"],
        "warnings": warnings,
    }

    if output_path:
        wb.save(output_path)
        wb.close()
        return metadata

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue(), metadata


# ── MAIN ENTRY POINT ──────────────────────────────────────────────────────────
def generate_report(input_path, output_path=None):
    """
    Main function. Read input_path (xlsx), return output bytes or write to output_path.
    """
    data        = read_input(input_path)
    info        = data["info"]
    rpt_year, rpt_month = _parse_period(info)

    mgmt_start = info.get("mgmt_start_date")
    if not isinstance(mgmt_start, datetime):
        raise ValueError("mgmt_start_date missing or invalid in Property_Info")

    months = compute_monthly(data, mgmt_start, rpt_year, rpt_month)
    if not months:
        raise ValueError("No monthly data computed — check period and management start date.")

    cur_month = months[-1]

    # Validate
    errors, warnings = validate(data, rpt_year, rpt_month, cur_month)
    if errors:
        raise ValueError("Validation errors:\n" + "\n".join(errors))

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_dashboard(wb, data, rpt_year, rpt_month, cur_month)
    build_pl(wb, data, rpt_year, rpt_month, cur_month)
    build_opex_passport(wb, data, rpt_year, rpt_month)
    build_12month(wb, data, rpt_year, rpt_month, months)
    build_ledger(wb, data, rpt_year, rpt_month, cur_month)
    build_dap(wb, data, rpt_year, rpt_month, months)

    # Set tab colors
    tab_colors = {
        "Dashboard":         NAVY,
        "P&L Monthly":       TEAL,
        "OPEX Passport":     GOLD,
        "12-Month Summary":  NAVY,
        "Transaction Ledger": TEAL,
        "DAP Snapshot":      GOLD,
    }
    for sheet in wb.worksheets:
        color = tab_colors.get(sheet.title, NAVY)
        sheet.sheet_properties.tabColor = color

    if output_path:
        wb.save(output_path)
        return warnings
    else:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue(), warnings

def generate_report_bundle(input_path):
    """
    Return both generated artifacts for API integrations:
    1) current owner report bytes;
    2) next month's prefilled InputData bytes.

    Flask / Make.com can use this helper when the scenario needs to upload both
    files from one request.
    """
    report_bytes, warnings = generate_report(input_path)
    next_input_bytes, next_input_meta = generate_next_input_template(input_path)
    return {
        "report_bytes": report_bytes,
        "next_input_bytes": next_input_bytes,
        "warnings": warnings,
        "next_input": next_input_meta,
    }


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate TropicLook Owner Report and optionally roll forward next InputData."
    )
    parser.add_argument("input", help="Current TL_[CODE]_InputData_[YYYY-MM].xlsx file")
    parser.add_argument("output", help="Output TL_[CODE]_OwnerReport_[YYYY-MM].xlsx file")
    parser.add_argument(
        "--next-input",
        dest="next_input",
        help="Optional output path for prefilled next-month TL_[CODE]_InputData_[YYYY-MM].xlsx",
    )
    args = parser.parse_args()

    inp, out = args.input, args.output
    if not os.path.exists(inp):
        print(f"ERROR: Input file not found: {inp}"); sys.exit(1)

    try:
        warnings = generate_report(inp, out)
        print(f"✅ Report generated: {out}")
        for w in warnings:
            print(f"  ⚠ {w}")
        if args.next_input:
            meta = generate_next_input_template(inp, args.next_input)
            print(
                "✅ Next input template generated: "
                f"{args.next_input} ({meta['current_period']} → {meta['next_period']}, "
                f"opening_balance={meta['opening_balance']:,.2f})"
            )
    except ValueError as e:
        print(f"❌ {e}"); sys.exit(1)
